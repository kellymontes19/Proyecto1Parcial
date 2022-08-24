from django.contrib.auth.models import User, Group
from openpyxl import Workbook
from django.forms import modelform_factory
from django.http import HttpResponse
from django.shortcuts import render, get_object_or_404, redirect

# Create your views here.
from rest_framework import viewsets

from docente.forms import DocenteFormulario
from docente.models import Docente, Direccion, Telefono
#Apolinario Nathaly
#Barcia Karol
#Basurto Laura
#Bellolio Karla
#Montes Kelly
#Pilozo Noelia
#Pincay Mell
from docente.serializers import DocenteSerializer, DireccionSerializer, TelefonoSerializer


def detalle_docente(request, id):
    # docente = Docente.objects.get(pk=id)
    docente = get_object_or_404(Docente, pk=id)
    mensaje = {'docente':docente}
    return render(request, 'detalle.html', mensaje)

#DocenteFormulario= modelform_factory(Docente,exclude=[])
def nuevo_docente(request):
    if request.POST:
        formDocente = DocenteFormulario (request.POST)
        if formDocente.is_valid():
            formDocente.save()
            return redirect('inicio')
    else:
        formDocente = DocenteFormulario()

    formDocente = DocenteFormulario()
    return render(request, 'nuevo.html',{'formDocente':formDocente})

def modificar_docente(request, id):
    docente = get_object_or_404(Docente, pk=id)
    if request.method == 'POST':
       formDocente = DocenteFormulario(request.POST, instance=docente)
       if formDocente.is_valid():
           formDocente.save()
           return redirect('inicio')
    else:
        formDocente = DocenteFormulario(instance=docente)
    mensaje = {'formDocente': formDocente}
    return render(request, 'modificar.html', context=mensaje)

def eliminar_docente(request, id):
    docente = get_object_or_404(Docente, pk=id)
    if docente:
        docente.delete()
    return redirect('inicio')

def reporte_docente(request):
    docente = Docente.objects.order_by('apellido')
    wb = Workbook()
    ws = wb.active
    ws['B1'] = 'REPORTE DE DOCENTES'
    ws.merge_cells('B1:E1')
    ws['B3'] ='ID'
    ws['C3'] ='NOMBRE'
    ws['D3'] ='APELLIDO'
    ws['E3'] ='EMAIL'
    cont = 4
    for docente in docente:
        ws.cell(row=cont, column=2).value = docente.id
        ws.cell(row=cont, column=3).value = docente.nombre
        ws.cell(row=cont, column=4).value = docente.apellido
        ws.cell(row=cont, column=5).value = docente.email
        cont = cont + 1
    nombre_archivo = "ReporteDocente.xlsx"
    response = HttpResponse(content_type="aplication/ms-excel")
    contenido = "attachment;filename={0}".format(nombre_archivo)
    response["Content-Disposition"] = contenido
    wb.save(response)
    return response
#Grupo 5
#Apolinario Nathaly
#Barcia Karol
#Basurto Laura
#Bellolio Karla
#Montes Kelly
#Pilozo Noelia
#Pincay Mell

class DocenteViewSet(viewsets.ModelViewSet):
    """
    API endpoint that allows users to be viewed or edited.
    """
    queryset = Docente.objects.all()
    serializer_class = DocenteSerializer
    #permission_classes = [permissions.IsAuthenticated]

class DireccionViewSet(viewsets.ModelViewSet):
    """
    API endpoint that allows users to be viewed or edited.
    """
    queryset = Direccion.objects.all()
    serializer_class = DireccionSerializer
    #permission_classes = [permissions.IsAuthenticated]

class TelefonoViewSet(viewsets.ModelViewSet):
    """
    API endpoint that allows users to be viewed or edited.
    """
    queryset = Telefono.objects.all()
    serializer_class = TelefonoSerializer
    #permission_classes = [permissions.IsAuthenticated]